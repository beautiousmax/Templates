""" This program allows user input to create a custom report
template. """

from openpyxl import Workbook, drawing, style
from Tkinter import Frame, Tk, Label, Checkbutton, Button, Entry, IntVar, W, \
    StringVar, OptionMenu, BOTTOM, END
from re import split
from collections import OrderedDict
import datetime
from sys import argv
import xml.etree.ElementTree as ET
import string


class App(Tk):
    """ This makes an interactive GUI with Tkinter in order to customize your
    report"""
    def __init__(self, parent, files):
        """ The meat of the GUI """
        Tk.__init__(self, parent)
        self.parent = parent

        self.title("Template Maker")
        self.geometry('540x710+50+50')

        self.keep = 0
        self.widgets = {}
        masterful = []
        dat = []

        if len(files) == 0:
            blanket = {"location": "",
                       "weather": "",
                       "entered": "CG",
                       "stars": "",
                       "north": "",
                       "east": "",
                       "south": "",
                       "west": "",
                       "month": "",
                       "day": "",
                       "year": "",
                       "interval": "15 minute",
                       "time1": "",
                       "time2": "",
                       "time3": "",
                       "time4": "",
                       "time5": "",
                       "time6": "",
                       "time7": "",
                       "time8": "",
                       "cars": 0,
                       "total_vehicles": 0,
                       "buses": 0,
                       "mediums": 0,
                       "heavy_trucks": 0,
                       "bicycles": 0,
                       "peds": 0,
                       "peds_bicycles": 0,
                       "rtor": 0}
            masterful.append(blanket)
        else:
            for item in files:
                blanket = {"location": "",
                           "weather": "",
                           "entered": "CG",
                           "stars": "",
                           "cars": 0,
                           "interval": "15 minute",
                           "total_vehicles": 0,
                           "buses": 0,
                           "mediums": 0,
                           "heavy_trucks": 0,
                           "bicycles": 0,
                           "peds": 0,
                           "peds_bicycles": 0,
                           "rtor": 0}
                data = shamwow(item)
                dat.append(data)
                date = data[2].split("/")
                blanket["month"] = date[0]
                blanket["day"] = date[1]
                blanket["year"] = date[2]
                legs = data[1].split(" ")
                print legs
                gestalt1 = data[1]
                gestalt2 = data[1]
                if ("/", "and", "&", "@") in legs:
                    for o, word in enumerate(legs):
                        if word in ("/", "and", "&", "@"):
                            gestalt1 = legs[0:o]
                            gestalt2 = legs[o + 1:]
                    if data[3]["north"] is True:
                        blanket["north"] = " ".join(gestalt1)
                    if data[3]["south"] is True:
                        blanket["south"] = " ".join(gestalt1)
                    if data[3]["east"] is True:
                        blanket["east"] = " ".join(gestalt2)
                    if data[3]["west"] is True:
                        blanket["west"] = " ".join(gestalt2)
                else:
                    if data[3]["north"] is True:
                        blanket["north"] = "".join(gestalt1)
                    if data[3]["south"] is True:
                        blanket["south"] = "".join(gestalt1)
                    if data[3]["east"] is True:
                        blanket["east"] = "".join(gestalt2)
                    if data[3]["west"] is True:
                        blanket["west"] = "".join(gestalt2)

                try:
                    blanket["time1"] = data[4][0][0].strftime("%H:%M")
                    blanket["time2"] = data[4][0][1].strftime("%H:%M")
                except IndexError:
                    blanket["time1"] = ""
                    blanket["time2"] = ""

                try:
                    blanket["time3"] = data[4][1][0].strftime("%H:%M")
                    blanket["time4"] = data[4][1][1].strftime("%H:%M")
                except IndexError:
                    blanket["time3"] = ""
                    blanket["time4"] = ""

                try:
                    blanket["time5"] = data[4][2][0].strftime("%H:%M")
                    blanket["time6"] = data[4][2][1].strftime("%H:%M")
                except IndexError:
                    blanket["time5"] = ""
                    blanket["time6"] = ""

                try:
                    blanket["time7"] = data[4][3][0].strftime("%H:%M")
                    blanket["time8"] = data[4][3][1].strftime("%H:%M")
                except IndexError:
                    blanket["time7"] = ""
                    blanket["time8"] = ""

                for name in data[5]:
                    if name == "Lights":
                        blanket["cars"] = 1
                    elif name in ("Other Vehicles", "Trucks",
                                  "Articulated Trucks"):
                        blanket["heavy_trucks"] = 1
                    elif name == "Buses":
                        blanket["buses"] = 1
                    elif name == "Mediums":
                        blanket["mediums"] = 1
                    elif "All Vehicles" in name:
                        blanket["total_vehicles"] = 1
                    elif name == "Pedestrians" and not blanket["peds_bicycles"]:
                        blanket["peds"] = 1
                    elif name == "Bicycles on Crosswalk":
                        blanket["peds"] = 0
                        blanket["peds_bicycles"] = 1
                    elif name == "Bicycles on Road":
                        blanket["bicycles"] = 1

                if data[6] is True:
                    blanket["rtor"] = 1

                masterful.append(blanket)

            counter = Frame(self)
            counter.pack()
            self.indexer = Entry(counter, width=2)
            right = Button(counter, width=5, text=">",
                           command=lambda: self.change_fields(masterful,
                                                              "right", files))
            left = Button(counter, width=5, text="<",
                          command=lambda: self.change_fields(masterful,
                                                             "left", files))

            right.grid(column=3, row=0)
            self.indexer.grid(column=2, row=0)
            left.grid(column=1, row=0)

            self.indexer.insert(0, self.keep + 1)

        Label(text="Header Information:").pack(pady=10)

        header = Frame(self)
        header.pack()

        Label(header, text="North Leg").grid(column=1, row=1)
        north = Entry(header)
        north.grid(column=2, row=1)

        Label(header, text="East Leg").grid(column=1, row=2)
        east = Entry(header)
        east.grid(column=2, row=2)
        Label(header, text="South Leg").grid(column=1, row=3)
        south = Entry(header)
        south.grid(column=2, row=3)
        Label(header, text="West Leg").grid(column=1, row=4)
        west = Entry(header)
        west.grid(column=2, row=4)
        self.widgets["north"] = north
        self.widgets["east"] = east
        self.widgets["south"] = south
        self.widgets["west"] = west
        Label(header, text="County, State").grid(column=1, row=5)
        loc = Entry(header)
        loc.grid(column=2, row=5)
        self.widgets["location"] = loc
        Label(header, text="Count Date").grid(column=1, row=6)

        countdate = Frame(header)
        countdate.grid(column=2, row=6)
        month = Entry(countdate, width=2)
        month.grid(column=2, row=1)
        Label(countdate, text="/").grid(column=3, row=1)
        day = Entry(countdate, width=2)
        day.grid(column=4, row=1)
        Label(countdate, text="/").grid(column=5, row=1)
        year = Entry(countdate, width=4)
        year.grid(column=6, row=1)
        self.widgets["month"] = month
        self.widgets["day"] = day
        self.widgets["year"] = year
        Label(header, text="Weather").grid(column=1, row=7)
        weat = Entry(header)
        weat.grid(column=2, row=7)
        self.widgets["weather"] = weat
        Label(header, text="Entered by").grid(column=1, row=8)
        entered = Entry(header)
        entered.grid(column=2, row=8)
        self.widgets["entered"] = entered
        Label(header, text="Star Rating").grid(column=1, row=9)
        star = Entry(header)
        star.grid(column=2, row=9)
        self.widgets["stars"] = star
        Label(text="Time Frame:").pack(pady=10)
        times = Frame(self)
        times.pack()

        Label(times, text="Start:").grid(column=1, row=1)
        time1 = Entry(times)
        time1.grid(column=2, row=1)
        Label(times, text="End:").grid(column=3, row=1)
        time2 = Entry(times)
        time2.grid(column=4, row=1)

        Label(times, text="Start:").grid(column=1, row=2)
        time3 = Entry(times)
        time3.grid(column=2, row=2)
        Label(times, text="End:").grid(column=3, row=2)
        time4 = Entry(times)
        time4.grid(column=4, row=2)

        Label(times, text="Start:").grid(column=1, row=3)
        time5 = Entry(times)
        time5.grid(column=2, row=3)
        Label(times, text="End:").grid(column=3, row=3)
        time6 = Entry(times)
        time6.grid(column=4, row=3)

        Label(times, text="Start:").grid(column=1, row=4)
        time7 = Entry(times)
        time7.grid(column=2, row=4)
        Label(times, text="End:").grid(column=3, row=4)
        time8 = Entry(times)
        time8.grid(column=4, row=4)
        self.widgets["time1"] = time1
        self.widgets["time2"] = time2
        self.widgets["time3"] = time3
        self.widgets["time4"] = time4
        self.widgets["time5"] = time5
        self.widgets["time6"] = time6
        self.widgets["time7"] = time7
        self.widgets["time8"] = time8
        interval = Frame(self)
        interval.pack()
        Label(interval, text="Select Interval:").grid(column=1, row=1)
        option_int = ["5 minute", "15 minute", "1 hour"]
        var = StringVar(interval)
        var.set(option_int[1])
        option = OptionMenu(interval, var, *option_int)
        option.grid(column=2, row=1)
        self.widgets["interval"] = var

        Label(text="Select Class Types Included:").pack(pady=10)
        tubs = Frame(self)
        tubs.pack()

        total_vehicles = IntVar()
        cars = IntVar()
        buses = IntVar()
        mediums = IntVar()
        heavy_trucks = IntVar()
        bicycles = IntVar()
        pedestrians = IntVar()
        pedestrians_bicycles = IntVar()
        rtor = IntVar()
        self.widgets["total_vehicles"] = total_vehicles
        self.widgets["cars"] = cars
        self.widgets["buses"] = buses
        self.widgets["mediums"] = mediums
        self.widgets["heavy_trucks"] = heavy_trucks
        self.widgets["bicycles"] = bicycles
        self.widgets["peds"] = pedestrians
        self.widgets["peds_bicycles"] = pedestrians_bicycles
        self.widgets["rtor"] = rtor

        check1 = Checkbutton(tubs, text="All Vehicles", variable=total_vehicles)
        check1.grid(sticky=W)
        check2 = Checkbutton(tubs, text="Cars", variable=cars)
        check2.grid(sticky=W)
        check3 = Checkbutton(tubs, text="Buses", variable=buses)
        check3.grid(sticky=W)
        check4 = Checkbutton(tubs, text="Mediums", variable=mediums)
        check4.grid(sticky=W)
        check5 = Checkbutton(tubs, text="Heavy Trucks", variable=heavy_trucks)
        check5.grid(sticky=W)
        check6 = Checkbutton(tubs, text="Bicycles", variable=bicycles)
        check6.grid(sticky=W)
        check7 = Checkbutton(tubs, text="Pedestrians", variable=pedestrians)
        check7.grid(sticky=W)
        check8 = Checkbutton(tubs, text="Pedestrians and bicycles",
                             variable=pedestrians_bicycles)
        check8.grid(sticky=W)
        check9 = Checkbutton(tubs, text="RTOR", variable=rtor)
        check9.grid(sticky=W)

        create = Button(self, text="Create",
                        command=lambda: self.save_fields(masterful, dat, True))
        create.pack(side=BOTTOM, pady=15)

        for k, v in self.widgets.iteritems():
            try:
                v.insert(0, masterful[self.keep][k])
            except AttributeError:
                v.set(masterful[self.keep][k])

    def save_fields(self, keys, d, s):
        """ This saves the stuff in the gui for use later. """
        for key, info in self.widgets.iteritems():
            if key == "location":
                for r in xrange(0, len(keys)):
                    if keys[r]["location"] == '':
                        keys[r]["location"] = info.get()
            if key == "weather":
                for r in xrange(0, len(keys)):
                    if keys[r]["weather"] == '':
                        keys[r]["weather"] = info.get()
            else:
                keys[self.keep][key] = info.get()
        if s is True:
            submit(keys)

    def change_fields(self, keys, direction, files):
        self.save_fields(keys, False, None)
        if direction == "right":
            if self.keep == len(files) - 1:
                self.keep = 0
            else:
                self.keep += 1
            self.indexer.delete(0, END)
            self.indexer.insert(0, self.keep + 1)
        elif direction == "left":
            if self.keep == 0:
                self.keep = len(files) - 1
            else:
                self.keep -= 1
            self.indexer.delete(0, END)
            self.indexer.insert(0, self.keep + 1)

        for k, v in self.widgets.iteritems():
            try:
                v.delete(0, END)
                v.insert(0, keys[self.keep][k])
            except AttributeError:
                v.set(keys[self.keep][k])

boxed_goodies = []


def shamwow(filer):
    """ This function holds 12x its weight in liquid. Shamwow does
    everything. Throw it in the washer when you're done. """
    directional_flow = []
    tree = ET.parse(filer)
    root = tree.getroot()

    all_data = {}
    sheets = []
    for child in root:
        datum = []
        t = ""
        for key, value in child.attrib.iteritems():
            t = value
            sheets.append(t)
        for things in child.findall(
                ".//{urn:schemas-microsoft-com:office:spreadsheet}Data"):
            datum.append(things.text)
        if len(datum) > 0:
            all_data[t] = datum

    study = ""
    start_date = ""
    orientation = []
    golden = ["Southbound", "Westbound", "Northbound", "Eastbound"]

    organized_data = {}
    for sheet, val in all_data.iteritems():
        directions = []
        b_directions = []
        best = OrderedDict()
        reds = 0
        dirc1 = ''
        dirc2 = ''
        for index, item in enumerate(val):
            b = []
            if item == "Study Name":
                study = val[index + 1]
            if item == "Start Date":
                start_date = val[index + 1]
            if item == "Start Time" and val[index - 2] != "Start Date":
                dirc1 = index + 1
            if len(directions) == 0 and index > 7:
                try:
                    if val[index + 1] is not None and val[index + 1][-2:] in ('AM', 'PM') and dirc2 == '':
                        dirc2 = index + 1
                        for direction in val[dirc1:dirc2]:
                            directions.append(direction)
                            if direction == "Right on Red":
                                reds += 1
                except IndexError:
                    pass

            if item in golden:
                orientation.append(item)

            if item is not None and item[-2:] in ('AM', 'PM') and val[index + 1] != "Site Code":
                u = 0
                a = val[index + 1: index + len(directions) + 1]
                for ix, d in enumerate(directions):
                    if d in ("U-Turn", 'Peds CW'):
                        if orientation[u] != golden[u]:
                            if 'Peds CW' in directions:
                                b.insert(ix - 1, '')
                                b_directions.append('Ped')
                            else:
                                b.insert(ix - 1, '')
                                b.insert(ix - 1, '')
                                b.insert(ix - 1, '')
                                b.insert(ix - 1, '')

                                b_directions.extend(["Right",
                                                     "Thru",
                                                     "Left",
                                                     "U-Turn"])
                            u += 1
                        u += 1

                    if 'Peds CW' not in directions:
                        if d not in ('Right', 'Right on Red'):
                            b.append(a[ix])
                            b_directions.append(d)

                        elif d in ('Right', 'Right on Red'):
                                if d == 'Right on Red' and sheet != 'Totals':
                                    b.append(str(int(a[ix]) + int(a[ix - 1])))
                                    b_directions.append('Right')
                                elif d == 'Right on Red' and sheet == 'Totals':
                                    b.append(a[ix])
                                    b_directions.append('Right on Red')
                                elif d == 'Right' and directions[ix + 1] != 'Right on Red':
                                    b.append(a[ix])
                                    b_directions.append('Right')

                                if directions[ix + 1] not in ("Thru",
                                                              "Right on Red"):
                                    b.append('')
                                    b_directions.append('Thru')

                        if ix == 0 and d != 'Right':
                            b.append('')
                            b_directions.append('Right')

                        if 0 < ix < len(directions) - 2:
                            if d == 'U-Turn' and directions[ix + 1] != 'Right':
                                b.append('')
                                b_directions.append('Right')

                            elif d == 'Thru' and directions[ix + 1] != "Left":
                                b.append('')
                                b_directions.append('Left')

                            elif d == 'Left' and directions[ix + 1] != "U-Turn":
                                b.append('')
                                b_directions.append('U-Turn')

                        elif ix == len(directions) - 1:
                            if d != 'U-Turn':
                                b.append('')
                                b_directions.append('U-Turn')

                    elif 'Peds CW' in directions:
                        if d == 'Peds Combined':
                            if a[ix - 1] is None:
                                b.append('')
                            else:
                                b.append(str(int(a[ix - 1]) + int(a[ix - 2])))
                            if sheet == 'Bicycles on Crosswalk':
                                b_directions.append('Bike')
                            else:
                                b_directions.append('Ped')
                        elif d == 'Peds CCW' and 'Peds Combined' \
                                not in directions:
                            if a[ix] is None:
                                b.append('')
                            else:
                                b.append(str(int(a[ix]) + int(a[ix + 1])))
                            if sheet == 'Bicycles on Crosswalk':
                                b_directions.append('Bike')
                            else:
                                b_directions.append('Ped')

                best[item] = b

        directions = b_directions
        organized_data[sheet] = best
        directional_flow.append(directions)
    boxed_goodies.append((directional_flow, organized_data))

    times = []
    leg_configure = {'north': True, 'east': True, 'south': True, 'west': True}

    for tick, dit in organized_data[sheets[0]].iteritems():
        a = datetime.datetime.strptime(tick, '%I:%M %p')
        times.append(a)
        if orientation:
            if dit[0:4] == [None, None, None, None]:
                leg_configure["north"] = False
            if dit[4:8] == [None, None, None, None]:
                leg_configure["east"] = False
            if dit[8:12] == [None, None, None, None]:
                leg_configure["south"] = False
            if dit[12:16] == [None, None, None, None]:
                leg_configure["west"] = False

    start = ''
    time_frame_segments = []
    for i, time in enumerate(times):
        if start == '':
            start = time
        elif time == times[-1] or time + datetime.timedelta(minutes=15) < \
                times[i + 1]:
            end = time + datetime.timedelta(minutes=15)
            time_frame_segments.append((start, end))
            start = ''

    return organized_data, study, start_date, leg_configure, time_frame_segments, sheets, True if reds else False


def formatted(sheet, s, *args):
    """ This method easily and quickly applies the proper text formatting to
    the specified cell."""
    for c in args:
        sheet.cell(c).style.font.name = "Arial"
        sheet.cell(c).style.font.size = 26
        sheet.cell(c).style.font.bold = True if s != 2 else False
        if s not in (1, 5, 6):
            sheet.cell(c).style.alignment.horizontal = style.Alignment.\
                HORIZONTAL_CENTER
        if s == 1 or s == 2:
            # Intersection, date, weather
            # Body of Data, times in sidebar
            sheet.cell(c).style.font.size = 29
        elif s == 3:
            # north leg, south leg, east leg
            sheet.cell(c).style.font.size = 26
        elif s == 4:
            # right thru left u
            sheet.cell(c).style.font.size = 22.5
            sheet.cell(c).style.borders.bottom.border_style = style.Border.\
                BORDER_THIN
        elif s == 5:
            # A1
            sheet.cell(c).style.font.size = 42
            sheet.column_dimensions['A'].width = 37
        else:
            # on: street name
            sheet.cell(c).style.font.size = 26
            sheet.cell(c).style.alignment.horizontal = style.Alignment.\
                HORIZONTAL_LEFT


def format_rows(sheet, *args):
    """ Just like the font formats, this formats the size of each row it is
    applied to."""
    for r in args:
        sheet.row_dimensions[r].height = 50


def format_border(sheet, col):
    """ This method applies borders to create four vertical columns at the
    specified row."""
    rows = ('A', 'F', 'K', 'P', 'U', 'V')
    for place in rows:
        sheet.cell('%s%s' % (place, col)).style.borders.right.\
            border_style = style.Border.BORDER_THIN
        sheet.cell('A%s' % col).style.borders.left.border_style = \
            style.Border.BORDER_THIN


def long_border(sheet, col):
    """ This applies a long horizontal border on the specified row."""
    for i in string.ascii_uppercase[1:21]:
        sheet.cell('%s%s' % (i, col)).style.borders.bottom.border_style = \
            style.Border.BORDER_THIN


def format_edges(sheet, row):
    """ The document needs some edges. This creates a border around it. """
    sheet.cell('V%s' % row).style.borders.right.border_style = \
        style.Border.BORDER_THIN
    sheet.cell('A%s' % row).style.borders.left.border_style = \
        style.Border.BORDER_THIN


def little_edge(sheet, row):
    """ This makes a cute little border on the edge. """
    sheet.cell('V%s' % row).style.borders.right.border_style = \
        style.Border.BORDER_THIN
    sheet.cell('A%s' % row).style.borders.left.border_style = \
        style.Border.BORDER_THIN
    sheet.cell('A%s' % row).style.borders.right.border_style = \
        style.Border.BORDER_THIN
    sheet.cell('U%s' % row).style.borders.right.border_style = \
        style.Border.BORDER_THIN


def format_top(sheet):
    """ The document needs a border at the top. This makes one."""
    for spot in string.ascii_uppercase[:22]:
        sheet.cell('%s1' % spot).style.borders.top.border_style = \
            style.Border.BORDER_THIN


def strip_time(header_dict):
    """ This function accepts 'normal' time and military time, and makes
    everything military time to process it easier."""
    for i in header_dict:
        if "time" in i and header_dict[i] != '':
            fragments = split(":| ", header_dict[i])
            hr = fragments[0]
            minute = fragments[1]
            if len(fragments) > 2:
                if fragments[2].lower() == "pm":
                    hr = int(hr) + 12
            header_dict[i] = (int(hr), int(minute))


def checks(header_dict):
    """ These are little checks to make sure the information you input into
    the gui can be processed properly. """
    total_sheets = 0
    total_times = 0
    errors = ""
    for i in header_dict:
        if i in ('cars', "total_vehicles", "Buses", 'heavy_trucks',
                 "bicycles", "peds", "peds_bicycles", "rtor"):
            total_sheets += header_dict[i]

    if total_sheets == 0:
        errors += "You need at least one class type selected. \n"

    if not header_dict["day"] and not header_dict["month"] and not \
            header_dict["year"]:
        errors += "You need to enter a date. Please use mm/dd/yyyy format. \n"
    elif header_dict["day"]:
        if len(header_dict["day"]) != 2 or not header_dict["day"].isdigit():
            errors += "The date looks funny. Please use mm/dd/yyyy format. \n"

        elif len(header_dict["month"]) != 2 or not \
                header_dict["month"].isdigit():
            errors += "The date looks funny. Please use mm/dd/yyyy format. \n"

        elif len(header_dict["year"]) != 4 or not header_dict["year"].isdigit():
            errors += "The date looks funny. Please use mm/dd/yyyy format. \n"

    if header_dict["stars"].isalpha() or len(header_dict["stars"]) > 1:
        errors += "Just one number for the star rating please."

    for i in header_dict:
        if "time" in i and header_dict[i] != '':
            total_times += 1
            if ":" not in header_dict[i]:
                errors += "Your time is missing a colon. \n"
            if not header_dict[i].replace(":", "").isdigit():
                if " " not in header_dict[i]:
                    errors += "If you are going to use am / pm, please put a " \
                              "space before it. \n"

    if not total_times:
        errors += "You need to add at least one interval. \n"

    if total_times % 2 != 0:
        errors += "Looks like you are missing part of an interval. \n"
    if not errors:
        return ''
    else:
        return errors


def pretty_time(a, b):
    """Here is where the time gets transformed from military time to be
    written to the worksheets."""
    if a[0] > 12 >= b[0]:
        better_a = (a[0]-12, a[1])
        return "%s:%s - %s:%s" % (better_a[0], str(better_a[1]).zfill(2), b[0],
                                  str(b[1]).zfill(2))
    elif b[0] > 12 >= a[0]:
        better_b = (b[0]-12, b[1])
        return "%s:%s - %s:%s" % (a[0], str(a[1]).zfill(2), better_b[0],
                                  str(better_b[1]).zfill(2))
    elif b[0] > 12 and a[0] > 12:
        better_a = (a[0]-12, a[1])
        better_b = (b[0]-12, b[1])
        return "%s:%s - %s:%s" % (better_a[0], str(better_a[1]).zfill(2),
                                  better_b[0], str(better_b[1]).zfill(2))
    else:
        return "%s:%s - %s:%s" % (a[0], str(a[1]).zfill(2), b[0],
                                  str(b[1]).zfill(2))


def refine_interval(x):
    """ This changes the string representing the interval to an integer that
    time_math can use."""
    if x == "15 minute":
        x = 15
    elif x == "1 hour":
        x = 60
    else:
        x = 5
    return x


def time_math(start, end, increment):
    """ This creates a list of times in the proper interval to be used in the
    side bar. """
    incremented_time = [start]
    hour = start[0]
    minute = start[1]
    if minute % increment == 0:
        if start == end:
            minute += increment
            if minute == 60:
                minute = 0
                hour += 1
            if hour == 24:
                hour = 0
            start = (hour, minute)
            incremented_time.append(start)

        while start != end:
            minute += increment
            if minute == 60:
                minute = 0
                hour += 1
            if hour == 24:
                hour = 0
            start = (hour, minute)
            incremented_time.append(start)

    return incremented_time


def make_filename(header_in):
    """ This makes finding a file name based on the intersection name. """
    if header_in["north"] == header_in["south"]:
        intersection_1 = header_in["north"]
    elif header_in["north"] == "" or header_in["south"] == "":
        if header_in["north"] == "":
            intersection_1 = header_in["south"]
        else:
            intersection_1 = header_in["north"]
    else:
        intersection_1 = "%s - %s" % (header_in["north"], header_in["south"])

    if header_in["east"] == header_in["west"]:
        intersection_2 = header_in["east"]
    elif header_in["east"] == "" or header_in["west"] == "":
        if header_in["east"] == "":
            intersection_2 = header_in["west"]
        else:
            intersection_2 = header_in["east"]
    else:
        intersection_2 = "%s - %s" % (header_in["east"], header_in["west"])

    d = "%s/%s/%s" % (header_in["month"], header_in["day"], header_in["year"])
    ok_date = datetime.datetime.strptime(d, "%m/%d/%Y")
    pretty_date = ok_date.strftime("%B %d, %Y")
    pretty_day = ok_date.strftime("%A")
    return (intersection_1, intersection_2, pretty_day, pretty_date)


def make_sheet(index, i, header_in, wb):
    """ This method is making the spread sheets, entering data in the
    appropriate cells. """
    nm = make_filename(header_in)

    w = wb.create_sheet()
    w.title = i
    w.page_setup.horizontalCentered = True
    w.page_setup.fitToPage = True
    w.page_setup.fitToHeight = 0
    w.page_setup.fitToWidth = 1
    w.page_setup.verticalCentered = True

    info = [('L2', "Counted by:"), ('L3', "Date:"), ('L4', "Weather:"),
            ('L5', "Entered by:"), ('C3', "Intersection of:"), ('C4', "and:"),
            ('C5', "Location:"), ('S5', 'Star Rating: ' + header_in["stars"])]

    for pair in info:
        w[pair[0]] = pair[1]
        formatted(w, 1, pair[0])
        w.cell(pair[0]).style.alignment.horizontal = style.Alignment.\
            HORIZONTAL_RIGHT

    info_data = [('D3', nm[0]), ('D4', nm[1]),
                 ('D5', header_in["location"]), ('M2', "VCU"),
                 ('M3', nm[3]), ('M4', header_in["weather"]),
                 ('M5', header_in["entered"]), ('R3', nm[2])]
    for data_pair in info_data:
        w[data_pair[0]] = data_pair[1]
        formatted(w, 1, data_pair[0])

    w['A9'] = "TIME"
    formatted(w, 3, 'A9')

    img = drawing.Image('J:\\PROGRAMS\\Templates\\TMC Formats\\logo.jpg')
    img.drawing.top = 5
    img.drawing.left = 2829
    w.add_image(img)

    format_rows(w, 2, 3, 4, 5)

    format_top(w)

    rows_to_edge = [1, 2, 3, 4, 5, 6]
    for edge in rows_to_edge:
        format_edges(w, edge)

    if i == "Pedestrians and Bicycles":
        side_bar1(index, w, header_in)
    elif i == "Pedestrians":
        side_bar2(index, w, header_in)
    else:
        side_bar(index, w, header_in)


def side_bar1(index, sheet, header_in):
    """ Pedestrians and bicycles """
    global tabs

    tabs = 11

    sheet['A1'] = "PEDESTRIAN AND BICYCLE OBSERVATIONS - SUMMARY"
    formatted(sheet, 5, 'A1')

    sheet.merge_cells('B7:K7')
    sheet['B7'] = "NORTH LEG"
    sheet.merge_cells('L7:U7')
    sheet['L7'] = "SOUTH LEG"

    sheet.merge_cells('B8:K8')
    sheet['B8'] = header_in["north"]
    sheet.merge_cells('L8:U8')
    sheet['L8'] = header_in["south"]

    sheet.merge_cells('B9:F9')
    sheet['B9'] = "Pedestrians"
    sheet.merge_cells('G9:K9')
    sheet['G9'] = "Bicycles"
    sheet.merge_cells('L9:P9')
    sheet['L9'] = "Pedestrians"
    sheet.merge_cells('Q9:U9')
    sheet['Q9'] = "Bicycles"
    formatted(sheet, 3, 'B7', 'L7', 'L8', 'B8', 'Q9', 'L9', 'G9', 'B9')
    format_border(sheet, 9)
    format_border(sheet, 10)

    needs_edges = [7, 8, 9, 10]
    for need in needs_edges:
        little_edge(sheet, need)

    sheet.cell('A6').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN
    long_border(sheet, 6)
    long_border(sheet, 8)
    sheet.cell('A10').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN
    long_border(sheet, 10)

    frame_options = [(header_in["time1"], header_in["time2"]),
                     (header_in["time3"], header_in["time4"]),
                     (header_in["time5"], header_in["time6"]),
                     (header_in["time7"], header_in["time8"])]

    for r in frame_options:
        if r[1]:
            if r[0][0] > 12:
                sheet['A%s' % tabs] = "PM"
            else:
                sheet['A%s' % tabs] = "AM"
            formatted(sheet, 2, 'A%s' % tabs)
            sheet.cell('A%s' % tabs).style.font.bold = True
            format_border(sheet, tabs)
            format_rows(sheet, tabs)
            tabs += 1

            list_interval = time_math(r[0], r[1], header_in["interval"])
            write_side_bar2(index, sheet, list_interval)
            sheet['A%s' % tabs] = "TOTALS"
            format_border(sheet, tabs)
            formatted(sheet, 3, 'A%s' % tabs)
            long_border(sheet, tabs - 1)
            long_border(sheet, tabs)
            sheet.cell('A%s' % (tabs - 1)).style.borders.bottom.\
                border_style = style.Border.BORDER_THIN
            sheet.cell('A%s' % tabs).style.borders.bottom.border_style = \
                style.Border.BORDER_THIN

            sum_start = (tabs - len(list_interval)) + 1
            sum_end = tabs - 1
            sheet.merge_cells('B%s:F%s' % (tabs, tabs))
            sheet['B%s' % tabs] = "=SUM(D%s:D%s)" % (sum_start, sum_end)
            sheet.merge_cells('G%s:K%s' % (tabs, tabs))
            sheet['G%s' % tabs] = "=SUM(I%s:I%s)" % (sum_start, sum_end)
            sheet.merge_cells('L%s:P%s' % (tabs, tabs))
            sheet['L%s' % tabs] = "=SUM(N%s:N%s)" % (sum_start, sum_end)
            sheet.merge_cells('Q%s:U%s' % (tabs, tabs))
            sheet['Q%s' % tabs] = "=SUM(S%s:S%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'B%s' % tabs, 'Q%s' % tabs, 'G%s' % tabs,
                      'L%s' % tabs)

            tabs += 1

    format_edges(sheet, tabs)

    tabs += 1
    format_edges(sheet, tabs)
    tabs += 1
    format_edges(sheet, tabs)
    tabs += 1
    format_edges(sheet, tabs)
    long_border(sheet, tabs)
    sheet.cell('A%s' % tabs).style.borders.bottom.border_style = style.\
        Border.BORDER_THIN
    tabs += 1

    sheet.merge_cells('B%s:K%s' % (tabs, tabs))
    sheet['B%s' % tabs] = "EAST LEG"
    sheet.merge_cells('L%s:U%s' % (tabs, tabs))
    sheet['L%s' % tabs] = "WEST LEG"
    formatted(sheet, 3, 'B%s' % tabs, 'L%s' % tabs)
    little_edge(sheet, tabs)

    tabs += 1

    long_border(sheet, tabs)
    sheet.merge_cells('B%s:K%s' % (tabs, tabs))
    sheet['B%s' % tabs] = header_in["east"]
    sheet.merge_cells('L%s:U%s' % (tabs, tabs))
    sheet['L%s' % tabs] = header_in["west"]
    formatted(sheet, 3, 'B%s' % tabs, 'L%s' % tabs)
    little_edge(sheet, tabs)

    tabs += 1
    sheet.merge_cells('B%s:F%s' % (tabs, tabs))
    sheet['B%s' % tabs] = "Pedestrians"
    sheet.merge_cells('G%s:K%s' % (tabs, tabs))
    sheet['G%s' % tabs] = "Bicycles"
    sheet.merge_cells('L%s:P%s' % (tabs, tabs))
    sheet['L%s' % tabs] = "Pedestrians"
    sheet.merge_cells('Q%s:U%s' % (tabs, tabs))
    sheet['Q%s' % tabs] = "Bicycles"
    formatted(sheet, 3, 'B%s' % tabs, 'G%s' % tabs, 'L%s' % tabs,
              'Q%s' % tabs)
    little_edge(sheet, tabs)
    format_border(sheet, tabs)
    tabs += 1
    format_border(sheet, tabs)
    sheet.cell('A%s' % tabs).style.borders.bottom.border_style = style.\
        Border.BORDER_THIN
    long_border(sheet, tabs)
    little_edge(sheet, tabs)

    tabs += 1
    frame_options = [(header_in["time1"], header_in["time2"]),
                     (header_in["time3"], header_in["time4"]),
                     (header_in["time5"], header_in["time6"]),
                     (header_in["time7"], header_in["time8"])]

    for r in frame_options:
        if r[1] != "":
            if r[0][0] > 12:
                sheet['A%s' % tabs] = "PM"
            else:
                sheet['A%s' % tabs] = "AM"

            formatted(sheet, 2, 'A%s' % tabs)
            sheet.cell('A%s' % tabs).style.font.bold = True
            format_border(sheet, tabs)
            format_rows(sheet, tabs)
            tabs += 1

            list_interval = time_math(r[0], r[1], header_in["interval"])
            write_side_bar2(index, sheet, list_interval, True)
            sheet['A%s' % tabs] = "TOTALS"
            format_border(sheet, tabs)
            formatted(sheet, 3, 'A%s' % tabs)
            long_border(sheet, tabs - 1)
            long_border(sheet, tabs)
            sheet.cell('A%s' % (tabs - 1)).style.borders.bottom.\
                border_style = style.Border.BORDER_THIN
            sheet.cell('A%s' % tabs).style.borders.bottom.border_style = \
                style.Border.BORDER_THIN

            sum_start = (tabs - len(list_interval)) + 1
            sum_end = tabs - 1
            sheet.merge_cells('B%s:F%s' % (tabs, tabs))
            sheet['B%s' % tabs] = "=SUM(D%s:D%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'B%s' % tabs)
            sheet.merge_cells('G%s:K%s' % (tabs, tabs))
            sheet['G%s' % tabs] = "=SUM(I%s:I%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'G%s' % tabs)
            sheet.merge_cells('L%s:P%s' % (tabs, tabs))
            sheet['L%s' % tabs] = "=SUM(N%s:N%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'L%s' % tabs)
            sheet.merge_cells('Q%s:U%s' % (tabs, tabs))
            sheet['Q%s' % tabs] = "=SUM(S%s:S%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'Q%s' % tabs)

            tabs += 1

    for thing in string.ascii_uppercase[1:22]:
        sheet.column_dimensions[thing].width = 19


def side_bar2(index, sheet, header_in):
    """ Just Pedestrians """
    global tabs

    tabs = 12

    sheet['A1'] = "PEDESTRIAN OBSERVATIONS - SUMMARY"
    formatted(sheet, 5, 'A1')

    sheet.merge_cells('B7:F7')
    sheet['B7'] = "NORTH LEG"
    sheet.merge_cells('G7:K7')
    sheet['G7'] = "SOUTH LEG"
    sheet.merge_cells('L7:P7')
    sheet['L7'] = "EAST LEG"
    sheet.merge_cells('Q7:U7')
    sheet['Q7'] = "WEST LEG"
    sheet.merge_cells('B8:F8')
    sheet['B8'] = header_in["north"]
    sheet.merge_cells('G8:K8')
    sheet['G8'] = header_in["south"]
    sheet.merge_cells('L8:P8')
    sheet['L8'] = header_in["east"]
    sheet.merge_cells('Q8:U8')
    sheet['Q8'] = header_in["west"]
    formatted(sheet, 3, 'B7', 'G7', 'L7', 'Q7', 'B8', 'G8', 'L8', 'Q8')

    borderer = [7, 8, 9, 10, 11]
    for place in borderer:
        format_border(sheet, place)

    sheet.cell('A10').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN
    sheet.cell('A11').style.borders.right.border_style = style.Border.\
        BORDER_THIN
    sheet.cell('A6').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN

    frame_options = [(header_in["time1"], header_in["time2"]),
                     (header_in["time3"], header_in["time4"]),
                     (header_in["time5"], header_in["time6"]),
                     (header_in["time7"], header_in["time8"])]

    for r in frame_options:
        if r[1] and r[0]:
            if r[0][0] > 12:
                sheet['A%s' % tabs] = "PM"
            else:
                sheet['A%s' % tabs] = "AM"
            formatted(sheet, 2, 'A%s' % tabs)
            format_border(sheet, tabs)
            sheet.cell('A%s' % tabs).style.font.bold = True
            format_rows(sheet, tabs)
            tabs += 1

            list_interval = time_math(r[0], r[1], header_in["interval"])
            write_side_bar2(index, sheet, list_interval)
            sheet['A%s' % tabs] = "TOTALS"
            format_border(sheet, tabs)
            formatted(sheet, 3, 'A%s' % tabs)
            sheet.cell('A%s' % tabs).style.borders.bottom.border_style = \
                style.Border.BORDER_THIN
            long_border(sheet, tabs)
            sheet.cell('A%s' % (tabs - 1)).style.borders.bottom.\
                border_style = style.Border.BORDER_THIN
            long_border(sheet, tabs - 1)

            sum_start = (tabs - len(list_interval)) + 1
            sum_end = tabs - 1
            sheet.merge_cells('B%s:F%s' % (tabs, tabs))
            sheet['B%s' % tabs] = "=SUM(D%s:D%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'B%s' % tabs)
            sheet.merge_cells('G%s:K%s' % (tabs, tabs))
            sheet['G%s' % tabs] = "=SUM(I%s:I%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'G%s' % tabs)
            sheet.merge_cells('L%s:P%s' % (tabs, tabs))
            sheet['L%s' % tabs] = "=SUM(N%s:N%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'L%s' % tabs)
            sheet.merge_cells('Q%s:U%s' % (tabs, tabs))
            sheet['Q%s' % tabs] = "=SUM(S%s:S%s)" % (sum_start, sum_end)
            formatted(sheet, 2, 'Q%s' % tabs)

            tabs += 1

    for thing in string.ascii_uppercase[1:22]:
        sheet.column_dimensions[thing].width = 19
        sheet.cell('%s6' % thing).style.borders.bottom.border_style = style.\
            Border.BORDER_THIN
        sheet.cell('%s8' % thing).style.borders.bottom.border_style = style.\
            Border.BORDER_THIN
        sheet.cell('%s10' % thing).style.borders.bottom.border_style = style.\
            Border.BORDER_THIN

    sheet.column_dimensions['V'].width = 19


def write_side_bar2(index, sheet, hrlist, second=False):
    """  This writes the body of the sheet for pedestrians and pedestrians
    with bicycles"""
    global tabs
    a = 0
    b = 1
    while b < len(hrlist):
        sheet['A%s' % tabs] = pretty_time(hrlist[a], hrlist[b])
        formatted(sheet, 2, 'A%s' % tabs)
        sheet.cell('A%s' % tabs).style.borders.right.border_style = style.\
            Border.BORDER_THIN
        format_border(sheet, tabs)
        format_rows(sheet, tabs)
        pedals = ['D%s' % tabs, 'I%s' % tabs, 'N%s' % tabs, 'S%s' % tabs]
        wonky_peds = [0, 2, 1, 3]
        if len(boxed_goodies) > 0:
            for x, i in enumerate(pedals):
                formatted(sheet, 2, i)
                try:
                    r = "{0}:{1} {2}".format(str(hrlist[a][0]) if hrlist[a][0] < 13 else str(hrlist[a][0] - 12), str(hrlist[a][1]).zfill(2), 'AM' if hrlist[a][0] < 12 else 'PM')
                    if 'Bicycles on Crosswalk' in boxed_goodies[index][1]:
                        wonky_bike_a = [('Pedestrians', 0), ('Bicycles on Crosswalk', 0), ('Pedestrians', 2), ('Bicycles on Crosswalk', 2)]
                        wonky_bike_b = [('Pedestrians', 1), ('Bicycles on Crosswalk', 1), ('Pedestrians', 3), ('Bicycles on Crosswalk', 3)]
                        if second is False:
                            sheet[i] = boxed_goodies[index][1][wonky_bike_a[x][0]][r][wonky_bike_a[x][1]]
                        else:
                            sheet[i] = boxed_goodies[index][1][wonky_bike_b[x][0]][r][wonky_bike_b[x][1]]
                    else:
                        sheet[i] = boxed_goodies[index][1]['Pedestrians'][r][wonky_peds[x]]
                except IndexError:
                    pass
        tabs += 1
        a += 1
        b += 1


def side_bar(index, sheet, header_in):
    """ This is for a 4 leg turning movement count - rtors / trucks, etc """
    global tabs
    tabs = 12

    sheet['A1'] = "%s TURNING MOVEMENT COUNT - SUMMARY" % sheet.title.upper()
    formatted(sheet, 5, 'A1')

    sheet.merge_cells('B7:F7')
    sheet['B7'] = "TRAFFIC FROM NORTH"
    sheet.merge_cells('G7:K7')
    sheet['G7'] = "TRAFFIC FROM SOUTH"
    sheet.merge_cells('L7:P7')
    sheet['L7'] = "TRAFFIC FROM EAST"
    sheet.merge_cells('Q7:U7')
    sheet['Q7'] = "TRAFFIC FROM WEST"
    sheet['B8'] = "on:"
    sheet['C8'] = header_in["north"]
    sheet['G8'] = "on:"
    sheet['H8'] = header_in["south"]
    sheet['L8'] = "on:"
    sheet['M8'] = header_in["east"]
    sheet['Q8'] = "on:"
    sheet['R8'] = header_in["west"]
    sheet['V7'] = "TOTAL"
    sheet['V8'] = "N + S"
    sheet['V9'] = "+"
    sheet['V10'] = "E + W"
    formatted(sheet, 6, 'C8', 'H8', 'M8', 'R8')
    formatted(sheet, 3, 'B7', 'G7', 'L7', 'Q7', 'B8',  'G8', 'L8', 'Q8', 'V7',
              'V8', 'V9', 'V10')

    rights = ['B10', 'G10', 'L10', 'Q10']
    for right in rights:
        sheet[right] = "RIGHT"
        formatted(sheet, 4, right)
    thrus = ['C10', 'H10', 'M10', 'R10']
    for thru in thrus:
        sheet[thru] = "THRU"
        formatted(sheet, 4, thru)
    lefts = ['D10', 'I10', 'N10', 'S10']
    for left in lefts:
        sheet[left] = "LEFT"
        formatted(sheet, 4, left)
    utns = ['E10', 'J10', 'O10', 'T10']
    for utn in utns:
        sheet[utn] = "U-TN"
        formatted(sheet, 4, utn)
    totals = ['F10', 'K10', 'P10', 'U10']
    for tot in totals:
        sheet[tot] = "TOTAL"
        formatted(sheet, 4, tot)

    borderer = [7, 8, 9, 10, 11]
    for place in borderer:
        format_border(sheet, place)

    sheet.cell('A10').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN
    sheet.cell('V10').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN
    sheet.cell('A6').style.borders.bottom.border_style = style.Border.\
        BORDER_THIN

    frame_options = [(header_in["time1"], header_in["time2"]),
                     (header_in["time3"], header_in["time4"]),
                     (header_in["time5"], header_in["time6"]),
                     (header_in["time7"], header_in["time8"])]

    for r in frame_options:
        if r[1]:
            if r[0][0] >= 12:
                sheet['A%s' % tabs] = "PM"
            else:
                sheet['A%s' % tabs] = "AM"

            formatted(sheet, 2, 'A%s' % tabs)
            sheet.cell('A%s' % tabs).style.font.bold = True
            format_border(sheet, tabs)
            format_rows(sheet, tabs)
            tabs += 1

            list_interval = time_math(r[0], r[1], header_in["interval"])

            write_side_bar(index, sheet, header_in, list_interval, False)

            sum_start = (tabs - len(list_interval)) + 1
            sum_end = tabs - 1

            sheet['A%s' % tabs] = "%s Hr Totals" % find_total_hours(
                list_interval, header_in)
            formatted(sheet, 2, 'A%s' % tabs)
            format_rows(sheet, tabs)
            sheet.cell('A%s' % tabs).style.borders.right.border_style = \
                style.Border.BORDER_THIN
            sheet.cell('A%s' % tabs).style.font.bold = True
            for column in string.ascii_uppercase[1:22]:
                sheet["%s%s" % (column, tabs)] = "=SUM(%s%s:%s%s)" % (
                    column, sum_start, column, sum_end)
                formatted(sheet, 2, "%s%s" % (column, tabs))
            format_border(sheet, tabs)
            tabs += 1

            if header_in["interval"] == 15:
                sheet['A%s' % tabs] = "1 Hr Totals"
                formatted(sheet, 2, 'A%s' % tabs)
                format_rows(sheet, tabs)
                sheet.cell('A%s' % tabs).style.font.bold = True
                sheet.cell('A%s' % tabs).style.borders.right.\
                    border_style = style.Border.BORDER_THIN
                format_border(sheet, tabs)
                tabs += 1
                write_side_bar(index, sheet, header_in, list_interval, True)

                if len(list_interval) > 39:
                    sheet['A%s' % tabs] = "PEAK HOUR"
                    formatted(sheet, 2, 'A%s' % tabs)
                    sheet.cell('A%s' % tabs).style.font.bold = True
                    format_border(sheet, tabs)
                    format_rows(sheet, tabs)
                    tabs += 1
                    formatted(sheet, 2, 'A%s' % tabs)
                    sheet.cell('A%s' % tabs).style.font.bold = True
                    format_border(sheet, tabs)
                    format_rows(sheet, tabs)
                    sheet.cell('A%s' % tabs).style.borders.top.\
                        border_style = style.Border.BORDER_DOUBLE
                    for column in string.ascii_uppercase[1:22]:
                        sheet.cell('%s%s' % (column, tabs)).style.borders.\
                            top.border_style = style.Border.BORDER_DOUBLE
                        formatted(sheet, 2, '%s%s' % (column, tabs))

                    tabs += 1

                    formatted(sheet, 2, 'A%s' % tabs)
                    sheet.cell('A%s' % tabs).style.font.bold = True
                    format_border(sheet, tabs)
                    sheet.cell('A%s' % tabs).style.borders.bottom.\
                        border_style = style.Border.BORDER_DOUBLE
                    sheet.cell('A%s' % tabs).style.borders.top.\
                        border_style = style.Border.BORDER_DOUBLE
                    format_rows(sheet, tabs)
                    for column in string.ascii_uppercase[1:22]:
                        sheet.cell('%s%s' % (column, tabs)).style.borders.\
                            bottom.border_style = style.Border.BORDER_DOUBLE
                        sheet.cell('%s%s' % (column, tabs)).style.borders.\
                            top.border_style = style.Border.BORDER_DOUBLE
                        formatted(sheet, 2, '%s%s' % (column, tabs))
                    tabs += 1

                else:
                    sheet['A%s' % tabs] = "PEAK HOUR"
                    formatted(sheet, 2, 'A%s' % tabs)
                    sheet.cell('A%s' % tabs).style.font.bold = True
                    format_border(sheet, tabs)

                    tabs += 1
                    formatted(sheet, 2, 'A%s' % tabs)
                    sheet.cell('A%s' % tabs).style.font.bold = True
                    format_border(sheet, tabs)
                    sheet.cell('A%s' % tabs).style.borders.bottom.\
                        border_style = style.Border.BORDER_DOUBLE
                    sheet.cell('A%s' % tabs).style.borders.top.\
                        border_style = style.Border.BORDER_DOUBLE
                    format_rows(sheet, tabs)

                    for column in string.ascii_uppercase[1:22]:
                        sheet.cell('%s%s' % (column, tabs)).style.borders.\
                            bottom.border_style = style.Border.BORDER_DOUBLE
                        sheet.cell('%s%s' % (column, tabs)).style.borders.\
                            top.border_style = style.Border.BORDER_DOUBLE
                        formatted(sheet, 2, '%s%s' % (column, tabs))
                    tabs += 1

    for thing in string.ascii_uppercase[1:22]:
        sheet.column_dimensions[thing].width = 19
        sheet.cell('%s6' % thing).style.borders.bottom.border_style = style.\
            Border.BORDER_THIN


def write_side_bar(index, sheet, sheets, hrlist, indicator):
    """ This does the heavy lifting for writing to the TMC side bar."""
    global tabs

    if indicator is True:
        a = 0
        b = 4
    else:
        a = 0
        b = 1

    teacup = ''
    if sheet.title == 'Vehicles':
        teacup = 'All Vehicles (no classificat...'
    elif sheet.title == "Heavy Trucks":
        if 'Articulated Trucks' in boxed_goodies[index][1]:
            teacup = 'Articulated Trucks'
        elif 'Trucks' in boxed_goodies[index][1]:
            teacup = "Trucks"
        else:
            teacup = 'Other Vehicles'
    elif sheet.title == "Buses":
        teacup = "Buses"
    elif sheet.title == "Cars":
        teacup = "Lights"
    elif sheet.title == "Mediums":
        teacup = "Mediums"
    elif sheet.title == "RTOR":
        teacup = "Totals"
    elif sheet.title == "Bicycles":
        teacup = "Bicycles on Road"
    elif sheet.title == "TOTALS":
        pass

    while b < len(hrlist):
        sheet["A%s" % tabs] = pretty_time(hrlist[a], hrlist[b])
        formatted(sheet, 2, "A%s" % tabs)
        sheet.cell("A%s" % tabs).style.borders.right.border_style = style.\
            Border.BORDER_THIN
        totals_columns = ["B", "C", "D", "E", "G", "H", "I", "J", "L", "M",
                          "N", "O", "Q", "R", "S", "T"]
        wonky_order = [0, 1, 2, 3, 8, 9, 10, 11, 4, 5, 6, 7, 12, 13, 14, 15]
        hour_start = (tabs - 1) - len(hrlist)
        hour_end = hour_start + 3
        for u, column in enumerate(totals_columns):
            if indicator is True:
                sheet["%s%s" % (column, tabs)] = "=SUM(%s%s:%s%s)" % (
                    column, hour_start, column, hour_end)
            else:
                if sheet.title == "TOTALS":
                    applicable_sheets = [i for i in sheets if i in
                                         ("cars", "heavy_trucks", "heavy_trucks")]

                    if applicable_sheets:
                        y = "="
                        for indexy, title_name in enumerate(applicable_sheets):
                            take = title_name.replace("_", " ").title()
                            if indexy < len(applicable_sheets) - 1:
                                y += "'{0}'!{1}{2}+".format(take, column, tabs)
                            else:
                                y += "'%s'!%s%s" % (take, column, tabs)
                        sheet['%s%s' % (column, tabs)] = y
                elif len(boxed_goodies) > 0:
                    if sheet.title == "RTOR":
                        r = "{0}:{1} {2}".format(str(hrlist[a][0]) if hrlist[a][0] < 13 else str(hrlist[a][0] - 12), str(hrlist[a][1]).zfill(2), 'AM' if hrlist[a][0] < 12 else 'PM')
                        if boxed_goodies[index][0][0][wonky_order[u]] == "Right on Red":
                            sheet['{0}{1}'.format(column, tabs)] = boxed_goodies[index][1][teacup][r][wonky_order[u]]

                    else:
                        try:
                            r = "{0}:{1} {2}".format(str(hrlist[a][0]) if hrlist[a][0] < 13 else str(hrlist[a][0] - 12), str(hrlist[a][1]).zfill(2), 'AM' if hrlist[a][0] < 12 else 'PM')
                            sheet['{0}{1}'.format(column, tabs)] = boxed_goodies[index][1][teacup][r][wonky_order[u]]
                        except IndexError or KeyError:
                            pass
            formatted(sheet, 2, "%s%s" % (column, tabs))

        sheet['F%s' % tabs] = "=SUM(B%s:E%s)" % (tabs, tabs)
        sheet['K%s' % tabs] = "=SUM(G%s:J%s)" % (tabs, tabs)
        sheet['P%s' % tabs] = "=SUM(L%s:O%s)" % (tabs, tabs)
        sheet['U%s' % tabs] = "=SUM(Q%s:T%s)" % (tabs, tabs)
        sheet['V%s' % tabs] = "=SUM(F{0}, K{0}, P{0}, U{0})".format(tabs)
        formatted(sheet, 2, 'F%s' % tabs, 'K%s' % tabs, 'P%s' % tabs,
                  'U%s' % tabs, 'V%s' % tabs)
        format_border(sheet, tabs)
        format_rows(sheet, tabs)

        tabs += 1
        a += 1
        b += 1


def trial(index, h_dict):
    """ This method is the springboard for writing to sheets, then saving the
    workbook. It also deletes the pesky default sheet that automatically comes
    with openpyxl workbooks."""
    wb = Workbook()
    titles = OrderedDict()
    titles["total_vehicles"] = "Vehicles" if h_dict["total_vehicles"] else ""
    titles["cars"] = "Cars" if h_dict["cars"] else ""
    titles["buses"] = "Buses" if h_dict["buses"] else ""
    titles["mediums"] = "Mediums" if h_dict["mediums"] else ""
    titles["heavy_trucks"] = "Heavy Trucks" if h_dict["heavy_trucks"] else ""
    titles["rtor"] = "RTOR" if h_dict["rtor"] else ""
    titles["bicycles"] = "Bicycles" if h_dict["bicycles"] else ""
    titles["peds"] = "Pedestrians" if h_dict["peds"] else ""
    titles["peds_bicycles"] = "Pedestrians and Bicycles" if h_dict["peds_bicycles"] else ""

    for i in titles:
        if titles[i]:
            make_sheet(index, titles[i], h_dict, wb)
    wb.remove_sheet(wb.worksheets[0])
    if titles["total_vehicles"] == "":
        make_sheet(index, "TOTALS", h_dict, wb)

    nm = make_filename(h_dict)
    filname = "%s %s & %s - %s" % (2014, nm[0], nm[1], nm[2])
    wb.save("%s.xlsx" % filname)


def find_total_hours(list_of_intervals, h_dict):
    """ Quickly find the total hours for the sidebar """
    if h_dict["interval"] == 15:
        hourable = 4
    elif h_dict["interval"] == 60:
        hourable = 1
    else:
        hourable = 12

    hour_total = (len(list_of_intervals)-1)/hourable
    return hour_total


def submit(header):
    """ This is what occurs when the 'create' button is clicked:
    All the entered data is pulled from the gui.
    The data is checked to make sure it makes some sense, the time is checked
    that it is in an ok format, etc.
    Excel spreadsheets are created using this information.
    """
    for index, item in enumerate(header):
        warning = checks(header[index])
        if warning == '':
            strip_time(header[index])
            header[index]["interval"] = refine_interval(header[index]["interval"])
            trial(index, header[index])
        else:
            erroneous = Tk()
            erroneous.title("Error Message")
            erroneous.geometry("390x200+100+100")

            Label(erroneous, text=warning).pack(pady=20)

            erroneous.mainloop()

if __name__ == "__main__":
    if len(argv) > 1:
        winn = App(None, argv[1:])
        winn.mainloop()
    else:
        win = App(None, [])
        win.mainloop()
